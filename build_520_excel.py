import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- 1. 读取原始数据 ----
with open('520_raw_result.json') as f:
    content = f.read()
start = content.find('[')
end = content.rfind(']') + 1
data = json.loads(content[start:end])
df = pd.DataFrame(data)

# 数值列转float
num_cols = ['order_cnt','actual_total_price','if_view_cnt_m_uv','if_ord_cnt_gmv_uv',
            'feishi_order_cnt','feishi_actual_total_price','feishi_if_view_cnt_m','feishi_if_ord_cnt_gmv',
            'ss_view_uv','ss_ord_uv']
for c in num_cols:
    df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)

# ---- 2. 拆分大促期 vs 基期 ----
# 大促期：20250519-20250520（2天）
# 基期：20250512-20250518（7天）
promo_days = 2
base_days = 7

promo = df[df['dt'].isin(['20250519','20250520'])].copy()
base = df[df['dt'].isin(['20250512','20250513','20250514','20250515','20250516','20250517','20250518'])].copy()

key_cols = ['prod_first_category_name','prod_second_category_name','prod_third_category_name']

promo_agg = promo.groupby(key_cols)[num_cols].sum().reset_index()
base_agg = base.groupby(key_cols)[num_cols].sum().reset_index()

# 合并
merged = promo_agg.merge(base_agg, on=key_cols, suffixes=('_promo','_base'), how='outer').fillna(0)

# ---- 3. 计算指标 ----
# 大促总量
merged['promo_gmv'] = merged['actual_total_price_promo']
merged['promo_order'] = merged['order_cnt_promo']
merged['promo_feishi_gmv'] = merged['feishi_actual_total_price_promo']
merged['promo_feishi_order'] = merged['feishi_order_cnt_promo']

# 基期总量
merged['base_gmv'] = merged['actual_total_price_base']
merged['base_order'] = merged['order_cnt_base']
merged['base_feishi_gmv'] = merged['feishi_actual_total_price_base']
merged['base_feishi_order'] = merged['feishi_order_cnt_base']
merged['base_ss_view_uv'] = merged['ss_view_uv_base']

# 均价（大促期）
merged['promo_avg_price'] = merged.apply(
    lambda r: r['promo_gmv'] / r['promo_order'] if r['promo_order'] > 0 else 0, axis=1)
merged['base_avg_price'] = merged.apply(
    lambda r: r['base_gmv'] / r['base_order'] if r['base_order'] > 0 else 0, axis=1)

# 爆发系数 = 大促日均 / 基期日均
def burst(promo_total, base_total, p_days=promo_days, b_days=base_days):
    base_daily = base_total / b_days if b_days > 0 else 0
    promo_daily = promo_total / p_days if p_days > 0 else 0
    if base_daily > 0:
        return promo_daily / base_daily
    return None

merged['burst_gmv'] = merged.apply(lambda r: burst(r['promo_gmv'], r['base_gmv']), axis=1)
merged['burst_order'] = merged.apply(lambda r: burst(r['promo_order'], r['base_order']), axis=1)
merged['burst_feishi_gmv'] = merged.apply(lambda r: burst(r['promo_feishi_gmv'], r['base_feishi_gmv']), axis=1)
merged['burst_feishi_order'] = merged.apply(lambda r: burst(r['promo_feishi_order'], r['base_feishi_order']), axis=1)
merged['price_change'] = merged.apply(
    lambda r: r['promo_avg_price'] / r['base_avg_price'] if r['base_avg_price'] > 0 else None, axis=1)

# ---- 4. 排序（按大促GMV降序）----
merged = merged.sort_values('promo_gmv', ascending=False).reset_index(drop=True)
merged['rank'] = merged.index + 1

# ---- 5. 写Excel ----
wb = openpyxl.Workbook()
ws = wb.active
ws.title = '520大促'

# 颜色定义
TITLE_BG = 'FF1F3864'   # 深蓝
PROMO_BG = 'FFBDD7EE'   # 浅蓝
BASE_BG  = 'FFE2EFDA'   # 浅绿
BURST_BG = 'FFFFF2CC'   # 浅黄
HEADER_FONT = Font(name='微软雅黑', bold=True, color='FFFFFFFF', size=10)
HEADER_FONT2 = Font(name='微软雅黑', bold=True, color='FF000000', size=9)
DATA_FONT = Font(name='微软雅黑', size=9)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center')

thin = Side(style='thin', color='FFB8B8B8')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def fill(hex_color):
    return PatternFill('solid', start_color=hex_color, end_color=hex_color)

# ---- 第1行：大标题 ----
ws.merge_cells('A1:U1')
ws['A1'] = '520大促 · 三级品类爆发系数明细（大促期：5/19-5/20，基期：5/12-5/18）'
ws['A1'].font = Font(name='微软雅黑', bold=True, color='FFFFFFFF', size=12)
ws['A1'].fill = fill('FF1F3864')
ws['A1'].alignment = CENTER

# ---- 第2行：分组标题 ----
# A-D: 基础信息, E-J: 大促, K-P: 基期, Q-U: 爆发系数
ws.merge_cells('A2:D2'); ws['A2'] = '基础信息'
ws.merge_cells('E2:J2'); ws['E2'] = f'520大促（共{promo_days}天）'
ws.merge_cells('K2:P2'); ws['K2'] = f'基期（20250512-20250518，共{base_days}天）'
ws.merge_cells('Q2:U2'); ws['Q2'] = '对比基期爆发系数（日均大促/日均基期）'

for cell, bg in [('A2','FFD6DCE4'),('E2','FF9DC3E6'),('K2','FF70AD47'),('Q2','FFFFC000')]:
    ws[cell].fill = fill(bg)
    ws[cell].font = Font(name='微软雅黑', bold=True, color='FF000000', size=10)
    ws[cell].alignment = CENTER

# ---- 第3行：列标题 ----
headers = ['rank','一级品类','二级品类','三级品类',
           '闪购消费GTV','闪购消费GTV(万)','闪购消费订单量','闪购消费订单量(万)','非食消费GTV','非食消费订单量',
           '闪购消费GTV','闪购消费订单量','非食消费GTV','非食消费订单量','搜索曝光UV','均价',
           '闪购消费GTV','闪购消费订单量','非食消费GTV','非食消费订单量','价格变化']
header_bgs = ['FFD6DCE4']*4 + ['FF9DC3E6']*6 + ['FF70AD47']*6 + ['FFFFC000']*5

for col_idx, (h, bg) in enumerate(zip(headers, header_bgs), start=1):
    cell = ws.cell(row=3, column=col_idx, value=h)
    cell.fill = fill(bg)
    cell.font = Font(name='微软雅黑', bold=True, color='FF000000', size=9)
    cell.alignment = CENTER
    cell.border = BORDER

# ---- 第4行起：数据 ----
data_cols = [
    'rank',
    'prod_first_category_name','prod_second_category_name','prod_third_category_name',
    'promo_gmv', None,  # GTV和GTV(万)
    'promo_order', None,  # 订单量和订单量(万)
    'promo_feishi_gmv','promo_feishi_order',
    'base_gmv','base_order','base_feishi_gmv','base_feishi_order','base_ss_view_uv','base_avg_price',
    'burst_gmv','burst_order','burst_feishi_gmv','burst_feishi_order','price_change'
]

row_bgs = ['FFD6DCE4']*4 + ['FFDEEAF1']*6 + ['FFE2EFDA']*6 + ['FFFFF2CC']*5

for row_idx, row_data in merged.iterrows():
    excel_row = row_idx + 4
    values = [
        int(row_data['rank']),
        row_data['prod_first_category_name'],
        row_data['prod_second_category_name'],
        row_data['prod_third_category_name'],
        row_data['promo_gmv'],
        row_data['promo_gmv'] / 10000,
        int(row_data['promo_order']),
        row_data['promo_order'] / 10000,
        row_data['promo_feishi_gmv'],
        int(row_data['promo_feishi_order']),
        row_data['base_gmv'],
        int(row_data['base_order']),
        row_data['base_feishi_gmv'],
        int(row_data['base_feishi_order']),
        int(row_data['base_ss_view_uv']),
        row_data['base_avg_price'],
        row_data['burst_gmv'],
        row_data['burst_order'],
        row_data['burst_feishi_gmv'],
        row_data['burst_feishi_order'],
        row_data['price_change'],
    ]
    for col_idx, (val, bg) in enumerate(zip(values, row_bgs), start=1):
        cell = ws.cell(row=excel_row, column=col_idx, value=val)
        cell.font = DATA_FONT
        cell.border = BORDER
        if col_idx <= 4:
            cell.alignment = LEFT
        else:
            cell.alignment = CENTER

# ---- 列宽 ----
col_widths = [6, 12, 14, 20, 14, 14, 12, 12, 14, 12, 14, 12, 14, 12, 12, 10, 10, 10, 10, 10, 10]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ---- 数字格式 ----
for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
    for cell in row:
        col = cell.column
        if col in [5, 9, 11, 13]:  # GMV列
            cell.number_format = '#,##0.00'
        elif col in [6, 8]:  # 万元列
            cell.number_format = '#,##0.0000'
        elif col == 16:  # 均价
            cell.number_format = '#,##0.00'
        elif col in [17, 18, 19, 20, 21]:  # 爆发系数
            cell.number_format = '0.0000'

# 冻结前3行
ws.freeze_panes = 'A4'

output_path = '/Users/lisicong/Desktop/kancong-daily/520大促三级品类明细.xlsx'
wb.save(output_path)
print(f'已保存: {output_path}')
print(f'数据行数: {len(merged)}')
